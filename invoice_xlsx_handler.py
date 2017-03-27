from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.worksheet.datavalidation import DataValidation
import internal_utils

COLUMN_HEADERS = {"A": "Date", "B": "Tech", "C": "TOW", "D": "Address", "E": "Class", "F": "Job Type", "G": "PSU",
                  "H": "Fast Fields\nreport\nreceived", "I": "NIA",
                  "J": "Complex Install Technician", "K": "Complex Install Materials", "L": "Invoiced", "M": "NOTES"}
COLUMN_HEADERS_FONT = Font(name='Calibri', size=12, bold=True)

SHEET_HEADER = "Launceston IT Invoicing Data"
SHEET_HEADER_FONT = Font(name='Calibri', size=18, bold=True)

DEFAULT_BORDER = Border(left=Side(border_style='thin'), right=Side(border_style='thin'), top=Side(border_style='thin'),
                        bottom=Side(border_style='thin'))

# maps job index to column letter
MAPPING = {0: "A", 1: "B", 2: "C", 3: "D", 4: "E", 5: "F", 6: "G", 7: "H", 8: "I", 9: "J", 10: "K", 11: "L",
           12: "M"}

CHECK_MARK = u'\u2713'

DV_B = DataValidation(type="list",
                      formula1='"Simon Timmers,Mike Richardson,Peter Richardson"',
                      allow_blank=True)

DV_E = DataValidation(type="list",
                      formula1='"1,2,3"',
                      allow_blank=True)

DV_F = DataValidation(type="list",
                      formula1='"End User - Underground,End User - Aerial,Battery Backup Install,FSD - Service Restoration,NTD Relocation"',
                      allow_blank=True)

DV_G = DataValidation(type="list",
                      formula1='"Battery,Battery Backup Install,Backup Standard,n / a"',
                      allow_blank=True)

DV_H = DataValidation(type="list",
                      formula1='"' + CHECK_MARK + ',N/A"',
                      allow_blank=True)

DV_L = DataValidation(type="list",
                      formula1='"' + CHECK_MARK + ',N/A"',
                      allow_blank=True)

DATA_VALIDATION_MAPPING = {"B": DV_B, "E": DV_E, "F": DV_F, "G": DV_G, "H": DV_H, "L": DV_L}


def write_jobs(jobs, worksheet, start_row=3):
    if isinstance(jobs, (list, tuple)):
        row = start_row
        for job in jobs:
            row += 1
            if isinstance(job, (list, tuple)):
                col = -1
                for cell in job:
                    col += 1
                    if col < len(MAPPING):
                        cell_text = MAPPING[col] + str(row)
                        worksheet[cell_text] = cell
                        worksheet[cell_text].border = DEFAULT_BORDER
                        dv = DATA_VALIDATION_MAPPING.get(MAPPING[col], None)
                        if dv:
                            dv.add(worksheet[cell_text])


def create_file_and_write_to(jobs, file_location="Invoice.xlsx"):
    wb = Workbook()

    # grab the active worksheet
    ws = wb.active

    # title
    ws.title = "Active"

    # sheet header
    ws.merge_cells('A1:L1')
    ws['A1'] = SHEET_HEADER
    ws['A1'].font = SHEET_HEADER_FONT
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 20

    # column headers
    for key, value in COLUMN_HEADERS.items():
        ws[key + '3'] = value
        ws[key + '3'].font = COLUMN_HEADERS_FONT
        ws[key + '3'].border = DEFAULT_BORDER
    ws.row_dimensions[3].height = 60

    # freeze rows
    ws.freeze_panes = 'A4'

    # data validation
    data_validation(ws)

    # jobs data
    write_jobs(jobs, ws)

    # resize columns
    internal_utils.resize_columns(ws, 2)

    # Save the file
    wb.save(file_location)


def data_validation(worksheet):
    worksheet.add_data_validation(DV_B)
    worksheet.add_data_validation(DV_E)
    worksheet.add_data_validation(DV_F)
    worksheet.add_data_validation(DV_G)
    worksheet.add_data_validation(DV_H)
    worksheet.add_data_validation(DV_L)


def add_to_file(jobs, file_location="invoice.xlsx"):
    # open an xlsx for reading
    wb = load_workbook(file_location)

    # grab the active worksheet
    ws = wb.active

    # data validation
    data_validation(ws)

    # jobs data
    write_jobs(jobs, ws, ws.max_row)

    # resize columns
    internal_utils.resize_columns(ws, 2)

    # save
    wb.save(file_location)


def read_jobs(worksheet, start_row=3):
    jobs = []
    for row in worksheet.iter_rows(min_row=start_row + 1):
        job = ["" for i in range(len(MAPPING))]
        for i, cell in enumerate(row):
            job[list(MAPPING.keys())[list(MAPPING.values()).index(cell.column)]] = internal_utils.as_text(cell.value)
        jobs.append(job)
    return jobs


def read_from_file(file_location="invoice.xlsx"):
    # open an xlsx for reading
    wb = load_workbook(file_location)

    # grab the active worksheet
    ws = wb.active

    # read data
    return read_jobs(ws)
