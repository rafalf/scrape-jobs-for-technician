from openpyxl.utils import get_column_letter

def as_text(value): return str(value) if value is not None else ""

def resize_columns(worksheet, min_row = 1):
    column_widths = []
    for row in worksheet.iter_rows(min_row=min_row):
        for i, cell in enumerate(row):
            try:
                column_widths[i] = max(column_widths[i], len(as_text(cell.value)))
            except IndexError:
                column_widths.append(len(as_text(cell.value)))

    for i, column_width in enumerate(column_widths):
        worksheet.column_dimensions[get_column_letter(i + 1)].width = column_width