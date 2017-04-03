from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.worksheet.worksheet import Worksheet

import merged_cell_styling

START_ROW = 1
INTERVAL_ROW = 8
DEFAULT_WIDTH = 12
DEFAULT_HEIGHT = 17
DEFAULT_MARGIN_LEFT = 0.1968503937
DEFAULT_MARGIN_RIGHT = 0.1181102362
DEFAULT_MARGIN_TOP = 0.7874015748
DEFAULT_MARGIN_BOTTOM = 0

DEFAULT_FONT = Font(name='Calibri', size=12)

# maps job index to column letter
DEFAULT_MAP = {0: ['C', 0], 1: ['B', 1], 2: ['B', 2], 3: ['D', 2], 4: ['F', 1], 5: ['F', 2], 6: ['B', 3], 7: ['F', 3],
               8: ['B', 4], 9: ['D', 4], 10: ['F', 4], 11: ['H', 4], 12: ['B', 5], 13: ['F', 5], 14: ['H', 5],
               15: ['B', 6], 16: ['F', 6]}

# LABELS = {"Date:": ['B', 0], "Address": ['A', 1],
#           "FDH": ['E', 1], "Phone": ['A', 2],
#           "Name": ['C', 2], "FDH Address": ['E', 2],
#           "Time": ['A', 3], "Local Port": ['E', 3],
#           "TOW": ['A', 4], "NTD ID": ['C', 4],
#           "Splitter": ['E', 4], "Out": ['G', 4],
#           "PSU Type": ['A', 5], "Multiport ": ['E', 5],
#           "Port": ['G', 5], "NTD Serial": ['A', 6], "PSU Serial": ['E', 6]}

LABELS = {"Date:": ['B', 0], "Address": ['A', 1],
          "Local Port": ['E', 1], "Phone": ['A', 2],
          "Name": ['C', 2], "FDH Address": ['E', 2],
          "Time": ['A', 3], "FDH": ['E', 3],
          "TOW": ['A', 4], "NTD ID": ['C', 4],
          "Splitter": ['E', 4], "Out": ['G', 4],
          "PSU Type": ['A', 5], "Multiport ": ['E', 5],
          "Port": ['G', 5], "NTD Serial": ['A', 6], "PSU Serial": ['E', 6]}

MERGED_CELLS = [[['C', 0], ['D', 0]],
                [['F', 1], ['H', 1]],
                [['B', 1], ['D', 1]],
                [['F', 2], ['H', 2]],
                [['B', 3], ['D', 3]],
                [['F', 3], ['H', 3]],
                [['B', 5], ['D', 5]],
                [['B', 6], ['D', 6]],
                [['F', 6], ['H', 6]]
                ]

BORDERED_CELLS = [['A', 1], ['E', 1], ['A', 2], ['B', 2], ['C', 2], ['D', 2], ['E', 2], ['A', 3], ['E', 3], ['A', 4],
                  ['B', 4], ['C', 4], ['D', 4], ['E', 4], ['F', 4],
                  ['G', 4], ['H', 4], ['A', 5], ['E', 5], ['F', 5], ['G', 5], ['H', 5], ['A', 6], ['E', 6]]

DEFAULT_BORDER = Border(left=Side(border_style='thin'), right=Side(border_style='thin'), top=Side(border_style='thin'),
                        bottom=Side(border_style='thin'))

RIGHT_ALIGNMENT = Alignment(horizontal='right')


def write_labels(worksheet, start_row=START_ROW):
    for key, value in LABELS.items():
        cell = value[0] + str(value[1] + start_row)
        worksheet[cell] = key
        worksheet[cell].font = DEFAULT_FONT
        if key == 'Date:':
            worksheet[cell].alignment = RIGHT_ALIGNMENT


def merge_cells(worksheet, start_row=START_ROW):
    for group in MERGED_CELLS:
        cell_range = group[0][0] + str(group[0][1] + start_row) + ':' + group[1][0] + str(group[1][1] + start_row)
        worksheet.merge_cells(cell_range)


def write_custom_borders(worksheet, start_row):
    for group in BORDERED_CELLS:
        worksheet[group[0] + str(group[1] + start_row)].border = DEFAULT_BORDER


def reborder_merged_cells(worksheet, end_row):
    for row in range(worksheet.min_row, end_row, INTERVAL_ROW):
        for group in MERGED_CELLS:
            cell_range = group[0][0] + str(group[0][1] + row) + ':' + group[1][0] + str(group[1][1] + row)
            if (group[1][1] % INTERVAL_ROW) != 0:
                merged_cell_styling.style_range(worksheet, cell_range, DEFAULT_BORDER)


def write_jobs(jobs, worksheet, start_row=START_ROW):
    if isinstance(jobs, (list, tuple)):
        row = start_row
        for job in jobs:
            write_labels(worksheet, row)
            merge_cells(worksheet, row)
            write_custom_borders(worksheet, row)
            if isinstance(job, (list, tuple)):
                for key, value in DEFAULT_MAP.items():
                    if key < len(job):
                        row_val = value[1] + row
                        cell_text = value[0] + str(row_val)
                        worksheet[cell_text] = job[key]
                        worksheet[cell_text].font = DEFAULT_FONT
                        worksheet.row_dimensions[row_val].height = DEFAULT_HEIGHT
            row += INTERVAL_ROW
        if jobs:
            reborder_merged_cells(worksheet, row)


def resize_columns(worksheet):
    worksheet.column_dimensions['A'].width = DEFAULT_WIDTH - 1.5 - 0.4
    worksheet.column_dimensions['B'].width = DEFAULT_WIDTH + 7 - 0.4
    worksheet.column_dimensions['C'].width = DEFAULT_WIDTH - 4.5 - 0.4
    worksheet.column_dimensions['D'].width = DEFAULT_WIDTH + 10 - 0.4
    worksheet.column_dimensions['E'].width = DEFAULT_WIDTH + 1 - 0.4
    worksheet.column_dimensions['F'].width = DEFAULT_WIDTH - 0.4
    worksheet.column_dimensions['G'].width = DEFAULT_WIDTH - 4 - 0.4
    worksheet.column_dimensions['H'].width = DEFAULT_WIDTH - 0.4

    worksheet.page_margins.left = DEFAULT_MARGIN_LEFT
    worksheet.page_margins.right = DEFAULT_MARGIN_RIGHT
    worksheet.page_margins.top = DEFAULT_MARGIN_TOP
    worksheet.page_margins.bottom = DEFAULT_MARGIN_BOTTOM


def write_jobs_and_save(workbook, jobs, file_location="jobs.xlsx"):
    ws = workbook.active

    next_row = START_ROW if ws.max_row == START_ROW else ws.max_row + START_ROW + 1
    write_jobs(jobs, ws, next_row)

    workbook.save(file_location)


def create_file_and_write_to(jobs, file_location="jobs.xlsx"):
    wb = Workbook()

    ws = wb.active

    ws.page_setup.paperSize = Worksheet.PAPERSIZE_A4
    ws.page_setup.fitToWidth = True

    resize_columns(ws)

    write_jobs_and_save(wb, jobs, file_location)


def add_to_file(jobs, file_location="jobs.xlsx"):
    wb = load_workbook(file_location)

    write_jobs_and_save(wb, jobs, file_location)
